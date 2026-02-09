import json
from io import BytesIO
from typing import Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .db import get_job, list_resumes, loads_modules
from .settings import DEFAULT_MODULES

MODULE_LABELS = {
    "SkillMatch": "技能匹配",
    "ProjectExperience": "项目经验",
    "YearsExperience": "工作年限",
    "Education": "教育背景",
    "Collaboration": "协作沟通",
    "Stability": "稳定性",
}

BASE_HEADERS_ZH = ["姓名", "联系方式", "邮箱", "总分", "总评"]


def _module_index(modules: List[Dict[str, str]]) -> Dict[str, Dict[str, str]]:
    return {m.get("name", ""): m for m in modules if isinstance(m, dict)}


def _load_profile_modules(job_id: str) -> List[Dict[str, str]]:
    job = get_job(job_id)
    if job and job.get("profile_json"):
        try:
            data = json.loads(job.get("profile_json") or "{}")
            modules = data.get("modules") if isinstance(data, dict) else None
            if isinstance(modules, list) and modules:
                return modules
        except json.JSONDecodeError:
            pass
    return [{"name": m.get("name", ""), "label": m.get("name", "")} for m in DEFAULT_MODULES]


def _module_label(module: Dict[str, str]) -> str:
    name = module.get("name", "")
    label = module.get("label") or ""
    if label and label != name:
        return label
    return MODULE_LABELS.get(name, label or name) or name


def build_rows(job_id: str, use_chinese: bool = True) -> Tuple[List[str], List[List[str]]]:
    profile_modules = _load_profile_modules(job_id)
    module_names = [m.get("name", "") for m in profile_modules if m.get("name")]
    if use_chinese:
        headers = BASE_HEADERS_ZH[:]
        for module in profile_modules:
            name = module.get("name", "")
            if not name:
                continue
            label = _module_label(module)
            headers.append(f"{label}分数")
            headers.append(f"{label}评价")
    else:
        headers = ["Name", "Phone", "Email", "TotalScore", "Summary"]
        for name in module_names:
            headers.append(f"{name}_Score")
            headers.append(f"{name}_Comment")

    resumes = list_resumes(job_id)
    rows = []
    for r in resumes:
        modules = loads_modules(r.get("modules_json"))
        module_map = _module_index(modules)
        total = r.get("total_score")
        row = [
            r.get("name") or "",
            r.get("phone") or "",
            r.get("email") or "",
            "" if total is None else f"{total:.1f}",
            r.get("summary") or "",
        ]
        for name in module_names:
            item = module_map.get(name, {})
            score = item.get("score")
            comment = item.get("comment") or ""
            if score is None:
                row.extend(["", comment])
            else:
                row.extend([f"{float(score):.1f}", comment])
        rows.append(row)

    rows.sort(key=lambda x: float(x[3]) if x[3] else -1, reverse=True)
    return headers, rows


def build_excel(job_id: str) -> BytesIO:
    headers, rows = build_rows(job_id, use_chinese=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    ws.append(headers)
    for row in rows:
        ws.append(row)

    for idx, header in enumerate(headers, start=1):
        col = get_column_letter(idx)
        ws.column_dimensions[col].width = max(12, len(header) + 2)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def build_markdown(job_id: str) -> str:
    headers, rows = build_rows(job_id, use_chinese=True)
    lines = []
    lines.append("| " + " | ".join(headers) + " |")
    lines.append("| " + " | ".join(["---"] * len(headers)) + " |")
    for row in rows:
        safe_row = [str(cell).replace("\n", " ") for cell in row]
        lines.append("| " + " | ".join(safe_row) + " |")
    return "\n".join(lines)
